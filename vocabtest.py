import openpyxl
import random
from openpyxl.styles import Border, Side, Alignment, Font

class vocabtest:
    def __init__(self, filename = "Book1.xlsx"):

        try:
            try : sheet = openpyxl.load_workbook(filename= filename)
            except: sheet = openpyxl.load_workbook(filename= filename + ".xlsx")
        except: sheet = openpyxl.load_workbook(filename="Book1.xlsx")
        sheet = sheet[sheet.sheetnames[0]]

        print("loading vocab...")

        unit = 0
        self.unitdicts = [[] for _ in range(50)]
        for i, _ in enumerate(sheet["A"]):
            for j, cell in enumerate(sheet[str(i+1)]):
                try : 
                    if cell.value.isascii() and cell.value.isalpha():
                        self.unitdicts[unit].append((cell.value, sheet[i+1][j+1].value))
                    else:
                        ix = cell.value.lower().find("unit") 
                        if ix != -1: unit = int(cell.value[ix+4:])
                except: pass
        self.tests = [] 
        print("done")

    #def gentest(self, units)

    def gentest(self, units:list, num:int, title: str):
        totdict = []
        for i in units:
            totdict += self.unitdicts[i]
        if num > len(totdict): 
            raise ValueError(f"Word count {num} exceed the total number of words, {len(totdict)}.") 
        random.shuffle(totdict)
        self.tests.append( (units, totdict[:num], title) )
        

    def puttest(self, tofile = False, filename = "./output.xlsx"):
        if tofile: 
            outputbook = openpyxl.Workbook()
            outputbook.save(filename= filename)

            # change the layout of a cell to make it printable
            # All borders, heiti 12, center alignment
            thinside = Side(style='thin', color = '000000')
            allborder = Border(left= thinside, right= thinside, top= thinside, bottom= thinside)
            heiti12 = Font(name= 'HeiTi', size= 12, bold= False, color= '000000')
            aligncc = Alignment(horizontal='center', vertical='center')
            def formatcell(cell):
                cell.border, cell.font, cell.alignment = allborder, heiti12, aligncc


        for test in self.tests:
            if tofile:

                # Create the Answer and Question Sheets
                answ_sheet = f"{str(test[2])} Answer".replace('[',' ').replace(']',' ')
                ques_sheet = f"{str(test[2])} Question".replace('[',' ').replace(']',' ')
                answ_sheet = outputbook.create_sheet(answ_sheet)
                ques_sheet = outputbook.create_sheet(ques_sheet)

                # Make the title
                for some_sheet in [answ_sheet, ques_sheet]:
                    some_sheet.row_dimensions[1].height = 30
                    for r1cell in some_sheet['A1:D1'][0]: # ((A1,B1,C1,D1),) (?)
                        formatcell(r1cell)
                    some_sheet.merge_cells('A1:D1')
                    some_sheet.cell(1, 1, test[2])

                for row in "ABCD":
                    answ_sheet.column_dimensions[row].width = 20
                    ques_sheet.column_dimensions[row].width = 20
            for i, (ur, ul) in enumerate(test[1]):
                if tofile: 
                    answ_sheet.row_dimensions[i//2+2].height = 26
                    formatcell(answ_sheet.cell(i//2+2, (i%2)*2+1, ul))
                    formatcell(answ_sheet.cell(i//2+2, (i%2)*2+2, ur))
                if ul == None: ul = ""
                if ur == None: ur = ""
                print(ul.ljust(7,' '), " \t", ur.ljust(15, ' '), end = "\t\n"[i%2])
        
            print("\n")

            for i, (_, ul) in enumerate(test[1]):
                if tofile: 
                    ques_sheet.row_dimensions[i//2+2].height = 26
                    formatcell(ques_sheet.cell(i//2+2, (i%2)*2+1, ul))
                    formatcell(ques_sheet.cell(i//2+2, (i%2)*2+2, None))
                    #output["Question"].cell(i//2+1, (i%2)*2+2, None)
                if ul == None: ul = ""
                print(ul.ljust(7,' '), " \t", " " * 15, end = "\t\n"[i%2])
        
        if tofile: 
            try: outputbook.remove(outputbook["Sheet"])
            except: print("Sheet 'Sheet' is already deleted. How wierd." )
            outputbook.save(filename= filename)
            outputbook.close()

    def numunit(self):
        return sum([1 if len(ux)!=0 else 0 for ux in self.unitdicts])
    def numword(self):
        return sum([len(ux) for ux in self.unitdicts])

if __name__ == "__main__":
    testi = vocabtest(input("input file name, default = Book1.xlsx"))
    print("input l, r, num")
    l, r, num = map(int, input().split())
    testi.gentest(list(range(l, r+1)), num=num, tofile=False)
    testi.puttest()
